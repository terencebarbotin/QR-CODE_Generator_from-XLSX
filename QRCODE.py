import qrcode
import qrcode.image.pil
import csv
import openpyxl
from openpyxl import load_workbook

def generate_qr_code(url, filename):
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    image = qr.make_image(image_factory=qrcode.image.pil.PilImage)
    image.save(filename)

filepath = "D:\Prog\Térence\Python\QR-CODE"

wb = load_workbook('Alumnis.xlsx')

sheet = wb.active

noms = []
prenoms = []
url_lk = []

for i in range(52):
    noms.append(sheet.cell(row=i+1, column=1).value)
    prenoms.append(sheet.cell(row=i+1, column=2).value)
    url_lk.append(sheet.cell(row=i+1, column=3).value)

    #print(str(i) + " : " + str(noms[i]) + " " + str(prenoms[i]) + " " + str(url_lk[i]))

print(url_lk)

for i in range(52):
    generate_qr_code(url_lk[i], str(noms[i]) + "_" + str(prenoms[i]) + ".png")

